"""
Demo seed script — run once before customer demo.

  cd leadpay/backend
  python3 seed_demo.py

What it does:
  1. Deletes all "Test Building *" and "Exp Building *" junk buildings (cascade).
  2. Creates 2 identical demo buildings (one for practice, one for the real demo).
  3. Seeds 15 apartments + 15 tenants per building.
  4. Seeds Jan, Feb, Mar bank statements + transactions (with confirmed matches).
  5. Generates demo_files/april_2026.xlsx and demo_files/may_2026.xlsx.
"""

import os
import sys
import uuid
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path

# Load backend .env
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import sqlalchemy as sa
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL not set in backend/.env")
    sys.exit(1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
Session = sessionmaker(bind=engine)


# ---------------------------------------------------------------------------
# Master data — same for both buildings
# ---------------------------------------------------------------------------

MONTHLY_PAYMENT = 800  # ₪800/month per apartment

TENANTS = [
    # (display_name, full_name, phone, apt_num, floor, bank_name)
    ("דוד לוי",         "דוד לוי",         "+972501234001", 1,  1, "לאומי"),
    ("מיכל כהן",        "מיכל כהן",        "+972521234002", 2,  1, "הפועלים"),
    ("יוסי אברהמי",     "יוסי אברהמי",     "+972541234003", 3,  1, "דיסקונט"),
    ("שרה גולדשטיין",   "שרה גולדשטיין",   "+972501234004", 4,  2, "לאומי"),
    ("אבי מזרחי",       "אבי מזרחי",       "+972521234005", 5,  2, "מזרחי"),
    ("רחל פרץ",         "רחל פרץ",         "+972541234006", 6,  2, "הפועלים"),
    ("מוחמד עמאר",      "מוחמד עמאר",      "+972501234007", 7,  3, "לאומי"),
    ("נועה שפירו",      "נועה שפירו",      "+972521234008", 8,  3, "דיסקונט"),
    ("עמי גרוס",        "עמי גרוס",        "+972541234009", 9,  3, "הפועלים"),
    ("לימור בן דוד",    "לימור בן דוד",    "+972501234010", 10, 4, "לאומי"),
    ("דנה ברוך",        "דנה ברוך",        "+972521234011", 11, 4, "מזרחי"),
    ("אורי נחמן",       "אורי נחמן",       "+972541234012", 12, 4, "הפועלים"),
    ("יעל שמש",         "יעל שמש",         "+972501234013", 13, 5, "לאומי"),
    ("גיל ריבקין",      "גיל ריבקין",      "+972521234014", 14, 5, "דיסקונט"),
    ("תמר עוזרי",       "תמר עוזרי",       "+972541234015", 15, 5, "הפועלים"),
]

# Payment amounts per tenant per month. None = did not pay, number = amount paid.
# Indexed 0-14 matching TENANTS list.
PAYMENTS = {
    1: {  # January 2026
        0: 800,   # דוד לוי       ✅
        1: 800,   # מיכל כהן      ✅
        2: 800,   # יוסי אברהמי   ✅
        3: 800,   # שרה גולדשטיין ✅
        4: None,  # אבי מזרחי     ❌
        5: 800,   # רחל פרץ       ✅
        6: 800,   # מוחמד עמאר    ✅
        7: 800,   # נועה שפירו    ✅
        8: 800,   # עמי גרוס      ✅
        9: 600,   # לימור בן דוד  ⚡ partial
        10: 800,  # דנה ברוך      ✅
        11: None, # אורי נחמן     ❌
        12: 800,  # יעל שמש       ✅
        13: 800,  # גיל ריבקין    ✅
        14: 800,  # תמר עוזרי     ✅
    },
    2: {  # February 2026
        0: 800,   # דוד לוי       ✅
        1: 500,   # מיכל כהן      ⚡ partial
        2: None,  # יוסי אברהמי   ❌
        3: 800,   # שרה גולדשטיין ✅
        4: None,  # אבי מזרחי     ❌
        5: 800,   # רחל פרץ       ✅
        6: 400,   # מוחמד עמאר    ⚡ partial
        7: 800,   # נועה שפירו    ✅
        8: None,  # עמי גרוס      ❌
        9: 800,   # לימור בן דוד  ✅
        10: 800,  # דנה ברוך      ✅
        11: None, # אורי נחמן     ❌
        12: 800,  # יעל שמש       ✅
        13: None, # גיל ריבקין    ❌
        14: 800,  # תמר עוזרי     ✅
    },
    3: {  # March 2026
        0: 800,   # דוד לוי       ✅
        1: 800,   # מיכל כהן      ✅
        2: 800,   # יוסי אברהמי   ✅
        3: 800,   # שרה גולדשטיין ✅
        4: 800,   # אבי מזרחי     ✅
        5: 800,   # רחל פרץ       ✅
        6: None,  # מוחמד עמאר    ❌
        7: 800,   # נועה שפירו    ✅
        8: None,  # עמי גרוס      ❌
        9: 800,   # לימור בן דוד  ✅
        10: 800,  # דנה ברוך      ✅
        11: None, # אורי נחמן     ❌
        12: 800,  # יעל שמש       ✅
        13: 800,  # גיל ריבקין    ✅
        14: 800,  # תמר עוזרי     ✅
    },
}

# Expenses per month: (description, debit_amount, building_category_name, legacy_category_string)
EXPENSES = {
    1: [  # January
        ("חברת החשמל", 1200, "חשמל", "routine_maintenance"),
        ("מקורות - מים", 300, "מים", "routine_maintenance"),
        ("חברת המעליות - תחזוקה", 600, "תיקונים", "technical_maintenance"),
    ],
    2: [  # February
        ("חברת החשמל", 1200, "חשמל", "routine_maintenance"),
        ("מקורות - מים", 300, "מים", "routine_maintenance"),
        ("חברת ניקיון ירוק", 400, "ניקיון", "routine_maintenance"),
    ],
    3: [  # March
        ("חברת החשמל", 1200, "חשמל", "routine_maintenance"),
        ("מקורות - מים", 300, "מים", "routine_maintenance"),
        ("חברת ניקיון ירוק", 400, "ניקיון", "routine_maintenance"),
    ],
}

MONTH_NAMES = {1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל", 5: "מאי"}


# ---------------------------------------------------------------------------
# Helper — make a date for a given month (pay on the 5th)
# ---------------------------------------------------------------------------

def pay_date(month: int, day: int = 5) -> datetime:
    return datetime(2026, month, day)


# ---------------------------------------------------------------------------
# Main seeding function
# ---------------------------------------------------------------------------

def seed(db, building_name: str, building_address: str, building_city: str) -> dict:
    """Seed one complete demo building. Returns a dict of IDs."""
    building_id = uuid.uuid4()
    print(f"\n  Creating building: {building_name} ({building_id})")

    db.execute(text("""
        INSERT INTO buildings (id, name, address, city, bank_account_number,
                               total_tenants, expected_monthly_payment, created_at, updated_at)
        VALUES (:id, :name, :address, :city, '12-345-678901',
                15, :emp, now(), now())
    """), {"id": str(building_id), "name": building_name, "address": building_address,
           "city": building_city, "emp": MONTHLY_PAYMENT})

    # Apartments
    apt_ids = {}
    for t in TENANTS:
        apt_num = t[2 + 1]  # apt_num is index 3
        floor   = t[2 + 2]  # floor   is index 4
        if apt_num not in apt_ids:
            apt_id = uuid.uuid4()
            apt_ids[apt_num] = apt_id
            db.execute(text("""
                INSERT INTO apartments (id, building_id, number, floor, expected_payment)
                VALUES (:id, :bid, :num, :floor, :ep)
            """), {"id": str(apt_id), "bid": str(building_id),
                   "num": apt_num, "floor": floor, "ep": MONTHLY_PAYMENT})

    # Tenants
    tenant_ids = {}
    for i, (name, full_name, phone, apt_num, floor, bank_name) in enumerate(TENANTS):
        tid = uuid.uuid4()
        tenant_ids[i] = tid
        db.execute(text("""
            INSERT INTO tenants (id, apartment_id, building_id, name, full_name, phone,
                                 ownership_type, bank_name, is_active, move_in_date,
                                 created_at, updated_at)
            VALUES (:id, :apt, :bid, :name, :full, :phone,
                    'שוכר', :bank, true, '2026-01-01',
                    now(), now())
        """), {"id": str(tid), "apt": str(apt_ids[apt_num]), "bid": str(building_id),
               "name": name, "full": full_name, "phone": phone, "bank": bank_name})

    # Default expense categories — must be inserted BEFORE expense allocations
    # so the statement loop can resolve category_id by name.
    for cat_name, cat_color in [
        ("ניקיון", "#4C72B0"), ("גינון", "#55A868"), ("חשמל", "#DD8452"),
        ("מים", "#8172B3"), ("תיקונים", "#C44E52"), ("אחר", "#937860"),
    ]:
        db.execute(text("""
            INSERT INTO expense_categories
                (id, building_id, name, color, is_default, is_active, created_at, updated_at)
            VALUES (:id, :bid, :name, :color, true, true, now(), now())
        """), {"id": str(uuid.uuid4()), "bid": str(building_id),
               "name": cat_name, "color": cat_color})

    # Bank statements + transactions for Jan, Feb, Mar
    for month in [1, 2, 3]:
        stmt_id = uuid.uuid4()
        period_label = f"{MONTH_NAMES[month]}_2026.xlsx"
        db.execute(text("""
            INSERT INTO bank_statements (id, building_id, upload_date, period_month, period_year,
                                         original_filename, raw_data)
            VALUES (:id, :bid, now(), :month, 2026, :fname, '{}')
        """), {"id": str(stmt_id), "bid": str(building_id),
               "month": month, "fname": period_label})

        balance = 50000.0  # fake running balance

        # Tenant payment transactions
        for i, (name, full_name, phone, apt_num, floor, bank_name) in enumerate(TENANTS):
            amount = PAYMENTS[month].get(i)
            if amount is None:
                continue
            tid = tenant_ids[i]
            txn_id = uuid.uuid4()
            description = f"{bank_name}    -  {full_name}"
            balance += amount
            db.execute(text("""
                INSERT INTO transactions (id, statement_id, activity_date, reference_number,
                    description, payer_name, credit_amount, debit_amount, balance,
                    transaction_type, matched_tenant_id, match_confidence, match_method,
                    is_confirmed, is_manual, created_at)
                VALUES (:id, :sid, :date, :ref, :desc, :payer, :credit, NULL, :bal,
                        'payment', :tenant, 0.99, 'exact', true, false, now())
            """), {
                "id": str(txn_id), "sid": str(stmt_id),
                "date": pay_date(month, day=5),
                "ref": str(1000 + i * 10 + month),
                "desc": description, "payer": full_name,
                "credit": amount, "bal": balance,
                "tenant": str(tid),
            })

            # Allocation for this payment
            alloc_id = uuid.uuid4()
            db.execute(text("""
                INSERT INTO transaction_allocations (id, transaction_id, tenant_id, label,
                    amount, period_month, period_year, created_at)
                VALUES (:id, :txn, :tenant, NULL, :amount, :month, 2026, now())
            """), {"id": str(alloc_id), "txn": str(txn_id), "tenant": str(tid),
                   "amount": amount, "month": month})

        # Expense transactions
        for desc, debit, cat_name, legacy_cat in EXPENSES.get(month, []):
            txn_id = uuid.uuid4()
            balance -= debit
            db.execute(text("""
                INSERT INTO transactions (id, statement_id, activity_date, reference_number,
                    description, payer_name, credit_amount, debit_amount, balance,
                    transaction_type, matched_tenant_id, match_confidence, match_method,
                    is_confirmed, is_manual, created_at)
                VALUES (:id, :sid, :date, :ref, :desc, NULL, NULL, :debit, :bal,
                        'transfer', NULL, NULL, NULL, true, false, now())
            """), {
                "id": str(txn_id), "sid": str(stmt_id),
                "date": pay_date(month, day=15),
                "ref": str(9000 + month * 10),
                "desc": desc, "debit": debit, "bal": balance,
            })

            # Resolve building's category_id by name (created earlier in this seed function).
            cat_row = db.execute(text("""
                SELECT id FROM expense_categories
                WHERE building_id = :bid AND name = :name
                LIMIT 1
            """), {"bid": str(building_id), "name": cat_name}).fetchone()
            cat_id = str(cat_row[0]) if cat_row else None

            alloc_id = uuid.uuid4()
            db.execute(text("""
                INSERT INTO transaction_allocations (id, transaction_id, tenant_id, label,
                    amount, period_month, period_year, category, category_id, created_at)
                VALUES (:id, :txn, NULL, :label, :amount, :month, 2026, :cat, :cat_id, now())
            """), {"id": str(alloc_id), "txn": str(txn_id), "label": desc,
                   "amount": debit, "month": month, "cat": legacy_cat, "cat_id": cat_id})

    db.commit()
    print(f"  Done: building_id={building_id}")
    return {"building_id": building_id, "tenant_ids": tenant_ids, "apt_ids": apt_ids}


# ---------------------------------------------------------------------------
# Excel generation helpers
# ---------------------------------------------------------------------------

EXCEL_COLS = ["תאריך פעילות", "אסמכתא", "תאור פעולה", "זכות", "חובה", "יתרה"]


def make_row(date_str: str, ref: str, description: str,
             credit=None, debit=None, balance: float = 0.0) -> list:
    return [date_str, ref, description,
            credit if credit else "",
            debit if debit else "",
            round(balance, 2)]


def write_excel(rows: list, filepath: Path) -> None:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "תנועות"

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(EXCEL_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    wb.save(filepath)
    print(f"  Saved: {filepath}")


# ---------------------------------------------------------------------------
# April 2026 Excel
# ---------------------------------------------------------------------------

def generate_april(output_dir: Path) -> None:
    print("\nGenerating April 2026 Excel...")
    rows = []
    balance = 65000.0
    ref = 4001

    def add_payment(name: str, amount: float, bank: str = "לאומי") -> None:
        nonlocal balance, ref
        balance += amount
        rows.append(make_row("05/04/26", str(ref), f"{bank}    -  {name}",
                              credit=amount, balance=balance))
        ref += 1

    def add_expense(desc: str, amount: float) -> None:
        nonlocal balance, ref
        balance -= amount
        rows.append(make_row("15/04/26", str(ref), desc,
                              debit=amount, balance=balance))
        ref += 1

    # 1. Auto-match tenants (₪800 each — exact names)
    add_payment("דוד לוי", 800)
    add_payment("שרה גולדשטיין", 800)
    add_payment("נועה שפירו", 800, bank="דיסקונט")
    add_payment("רחל פרץ", 800, bank="הפועלים")
    add_payment("דנה ברוך", 800, bank="מזרחי")
    add_payment("יעל שמש", 800)
    add_payment("תמר עוזרי", 800, bank="הפועלים")

    # 2. Manual match needed — abbreviated name
    add_payment("מ. כהן", 800)

    # 3. Multi-month catch-up — עמי גרוס owes Feb+Mar, pays ₪1,600
    add_payment("עמי גרוס", 1600, bank="הפועלים")

    # 4. Joint payment — two tenants in one transfer (manual split needed)
    balance += 1600
    rows.append(make_row("07/04/26", str(ref),
                          "דיסקונט    -  יוסי אברהמי וגיל ריבקין",
                          credit=1600, balance=balance))
    ref += 1

    # 5. Expenses (debits)
    add_expense("חברת החשמל", 1200)
    add_expense("מקורות - מים", 300)

    write_excel(rows, output_dir / "april_2026.xlsx")


# ---------------------------------------------------------------------------
# May 2026 Excel
# ---------------------------------------------------------------------------

def generate_may(output_dir: Path) -> None:
    print("\nGenerating May 2026 Excel...")
    rows = []
    balance = 68000.0
    ref = 5001

    def add_payment(name: str, amount: float, bank: str = "לאומי") -> None:
        nonlocal balance, ref
        balance += amount
        rows.append(make_row("05/05/26", str(ref), f"{bank}    -  {name}",
                              credit=amount, balance=balance))
        ref += 1

    def add_expense(desc: str, amount: float) -> None:
        nonlocal balance, ref
        balance -= amount
        rows.append(make_row("15/05/26", str(ref), desc,
                              debit=amount, balance=balance))
        ref += 1

    # 1. Auto-match ×6
    add_payment("לימור בן דוד", 800)
    add_payment("אבי מזרחי", 800, bank="מזרחי")
    add_payment("מוחמד עמאר", 800)
    add_payment("גיל ריבקין", 800, bank="דיסקונט")
    add_payment("יוסי אברהמי", 800, bank="דיסקונט")
    add_payment("שרה גולדשטיין", 800)

    # 2. Big catch-up — אורי נחמן pays 3 months of debt
    add_payment("אורי נחמן", 2400, bank="הפועלים")

    # 3. Unmatched mystery transaction (no tenant will match)
    balance += 800
    rows.append(make_row("08/05/26", str(ref),
                          "לאומי    -  א. כהן",
                          credit=800, balance=balance))
    ref += 1

    # 4. Expenses
    add_expense("חברת ניקיון ירוק", 400)
    add_expense("חברת החשמל", 1200)

    write_excel(rows, output_dir / "may_2026.xlsx")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("LeadPay Demo Seed Script")
    print("=" * 60)

    with Session() as db:
        # Step 1: Delete junk test/exp buildings (manual cascade order)
        print("\nStep 1: Deleting test/exp buildings...")
        junk = db.execute(text("""
            SELECT id FROM buildings
            WHERE name LIKE 'Test Building%' OR name LIKE 'Exp Building%'
        """)).fetchall()
        junk_ids = [str(r[0]) for r in junk]
        if junk_ids:
            ids_str = ",".join(f"'{i}'" for i in junk_ids)
            # Delete child records in dependency order
            db.execute(text(f"""
                DELETE FROM transaction_allocations
                WHERE transaction_id IN (
                    SELECT t.id FROM transactions t
                    JOIN bank_statements bs ON t.statement_id = bs.id
                    WHERE bs.building_id IN ({ids_str})
                )
            """))
            db.execute(text(f"""
                DELETE FROM transactions
                WHERE statement_id IN (
                    SELECT id FROM bank_statements WHERE building_id IN ({ids_str})
                )
            """))
            db.execute(text(f"DELETE FROM bank_statements WHERE building_id IN ({ids_str})"))
            db.execute(text(f"DELETE FROM name_mappings WHERE building_id IN ({ids_str})"))
            db.execute(text(f"DELETE FROM vendor_mappings WHERE building_id IN ({ids_str})"))
            db.execute(text(f"DELETE FROM messages WHERE building_id IN ({ids_str})"))
            # Tenants (with their allocations first)
            db.execute(text(f"""
                DELETE FROM transaction_allocations
                WHERE tenant_id IN (
                    SELECT id FROM tenants WHERE building_id IN ({ids_str})
                )
            """))
            db.execute(text(f"DELETE FROM tenants WHERE building_id IN ({ids_str})"))
            db.execute(text(f"DELETE FROM apartments WHERE building_id IN ({ids_str})"))
            db.execute(text(f"DELETE FROM buildings WHERE id IN ({ids_str})"))
            db.commit()
        print(f"  Deleted {len(junk_ids)} junk buildings.")

        # Step 2: Seed two demo buildings
        print("\nStep 2: Seeding demo buildings...")
        seed(db, "דמו - בניין לתרגול", "הגפן 12", "רמת גן")
        seed(db, "דמו - ללקוח", "הגפן 12", "רמת גן")

    # Step 3: Generate Excel files
    output_dir = Path(__file__).parent.parent / "demo_files"
    generate_april(output_dir)
    generate_may(output_dir)

    print("\n" + "=" * 60)
    print("Done! Summary:")
    print("  • Deleted all Test/Exp buildings")
    print("  • Created 2 demo buildings with 15 tenants each")
    print("  • Seeded Jan, Feb, Mar statements + transactions")
    print(f"  • Excel files in: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    main()
