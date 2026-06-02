"""
Tenants xlsx parser.

Parses the per-building "דיירים" (tenants) spreadsheet the user maintains by
hand. The sheet has a single fixed header row in Hebrew:

    דירה | קומה | סיווג | שם | טלפון | גובה תשלום
    (apt)  (floor) (class)  (name)  (phone)  (amount)

An "apartment block" starts with one owner (בעלים) row that carries the
apartment number, floor, and monthly payment amount. It may be followed by
zero or more renter (שוכר) rows whose דירה/קומה/גובה תשלום columns are blank
and whose שם/טלפון columns carry the renter's name and phone.

The parser is intentionally pure: it has no database dependency, raises
``ValueError`` on malformed input, and exposes a small set of normalization
helpers as testable internals (prefixed with ``_``).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

OWNER_CLASS = "בעלים"
RENTER_CLASS = "שוכר"

# Column indices (0-based) in the source sheet.
_COL_APARTMENT = 0
_COL_FLOOR = 1
_COL_CLASSIFICATION = 2
_COL_NAME = 3
_COL_PHONE = 4
_COL_AMOUNT = 5


# ─────────────────────────────────────────────────────────────────────
# Dataclasses
# ─────────────────────────────────────────────────────────────────────

@dataclass
class ParsedTenant:
    name: str                 # required, stripped, non-empty
    phone_raw: str | None     # normalized: comma-separated digit groups, or None
    classification: str       # exactly what the Excel said: "בעלים" or "שוכר"


@dataclass
class ParsedApartment:
    apartment_label: str               # "1", "מסחר 0", "דירת גן 2" — always a string
    floor_label: str | None            # "1", "קרקע", or None
    monthly_amount: Decimal | None     # Decimal("461") or None
    primary_tenant: ParsedTenant       # the owner (classification == OWNER_CLASS)
    renters: list[ParsedTenant] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────
# Normalization helpers (testable internals)
# ─────────────────────────────────────────────────────────────────────

def _normalize_apartment_label(raw: Any) -> str | None:
    """
    1.0 → '1', 1 → '1', 'מסחר 0' → 'מסחר 0', 1.5 → '1.5', '' → None, None → None.
    Floats with integer value lose the trailing '.0'.
    """
    return _normalize_label(raw)


def _normalize_floor_label(raw: Any) -> str | None:
    """Same contract as :func:`_normalize_apartment_label`."""
    return _normalize_label(raw)


def _normalize_label(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        # bools are ints in Python; not meaningful here.
        return None
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        # Drop trailing '.0' on whole-number floats; keep fractional part otherwise.
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    text = str(raw).strip()
    return text or None


def _normalize_phone(raw: Any) -> str | None:
    """
    Normalize a phone-column value.

    - Float (``544445201.0``) → ``'0544445201'`` (zero-pad to 10 digits since
      Excel strips the leading zero from numeric phones).
    - Int (``544445201``) → ``'0544445201'`` (same logic).
    - String with digits + separators → keep digit groups separated by
      ``', '`` when the cell contains multiple numbers; drop other punctuation
      within a number:
        - ``'054-6766257'`` → ``'0546766257'``
        - ``'0506246423, 0508584417'`` → ``'0506246423, 0508584417'``
        - ``'0506246423,0508584417'`` → ``'0506246423, 0508584417'`` (re-spaced)
    - ``None`` / empty / whitespace-only → ``None``.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        # Excel may have stripped leading zeros. Zero-pad to at least 10 digits
        # (an Israeli mobile is 10 digits including the leading 0).
        as_int = int(raw)
        if as_int < 0:
            return None
        digits = str(as_int)
        if len(digits) < 10:
            digits = digits.zfill(10)
        return digits

    text = str(raw).strip()
    if not text:
        return None

    # Split on commas → each part is one phone number; strip non-digits from
    # each part, then re-join with ', '.
    parts = [re.sub(r"\D", "", part) for part in text.split(",")]
    parts = [p for p in parts if p]
    if not parts:
        return None
    return ", ".join(parts)


def _parse_amount(raw: Any) -> Decimal | None:
    """
    ``'461 ₪'`` → ``Decimal('461')``. ``'461'`` → ``Decimal('461')``.
    ``461`` → ``Decimal('461')``. ``461.5`` → ``Decimal('461.5')``.
    ``None`` / empty → ``None``. Whitespace and the ₪ char are stripped.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return Decimal(raw)
    if isinstance(raw, float):
        # Use str() so that 461.5 → '461.5' not the noisy repr.
        return Decimal(str(raw))
    text = str(raw).replace("₪", "").strip()
    if not text:
        return None
    # Remove thousands separators if present.
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


# ─────────────────────────────────────────────────────────────────────
# Row-level helpers
# ─────────────────────────────────────────────────────────────────────

def _cell(row: tuple, index: int) -> Any:
    """Return ``row[index]`` if present, else ``None`` (defensive for short rows)."""
    if index < len(row):
        return row[index]
    return None


def _is_blank_row(row: tuple) -> bool:
    """A row is blank when every cell is None or whitespace-only."""
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def _normalize_text(raw: Any) -> str:
    """Coerce to a stripped string; return '' for None or blanks."""
    if raw is None:
        return ""
    return str(raw).strip()


# ─────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────

def parse_tenants_xlsx(path: Path) -> list[ParsedApartment]:
    """Parse the tenants xlsx file. Returns one ParsedApartment per apartment block.

    Raises ``ValueError`` if a renter row appears before any owner row, or if
    a row has an apartment number but a non-owner classification.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        # Stream rows; skip the header row (row 1).
        rows = ws.iter_rows(min_row=2, values_only=True)

        apartments: list[ParsedApartment] = []
        current: ParsedApartment | None = None

        for row_index, raw_row in enumerate(rows, start=2):
            # Trailing empty rows: openpyxl reports max_row including blanks.
            if _is_blank_row(raw_row):
                continue

            apartment_label = _normalize_apartment_label(_cell(raw_row, _COL_APARTMENT))
            floor_label = _normalize_floor_label(_cell(raw_row, _COL_FLOOR))
            classification = _normalize_text(_cell(raw_row, _COL_CLASSIFICATION))
            name = _normalize_text(_cell(raw_row, _COL_NAME))
            phone = _normalize_phone(_cell(raw_row, _COL_PHONE))
            amount = _parse_amount(_cell(raw_row, _COL_AMOUNT))

            # ── New apartment block: row carries an apartment number. ──
            if apartment_label is not None:
                if classification != OWNER_CLASS:
                    raise ValueError(
                        f"Row with apartment number must be classified as "
                        f"{OWNER_CLASS!r}, got {classification!r} at line {row_index}"
                    )
                if not name:
                    # Owner row with no name is meaningless; treat as malformed.
                    raise ValueError(
                        f"Owner row missing tenant name at line {row_index}"
                    )
                current = ParsedApartment(
                    apartment_label=apartment_label,
                    floor_label=floor_label,
                    monthly_amount=amount,
                    primary_tenant=ParsedTenant(
                        name=name,
                        phone_raw=phone,
                        classification=OWNER_CLASS,
                    ),
                )
                apartments.append(current)
                continue

            # ── Continuation row: belongs to the current apartment. ──
            if current is None:
                # No apartment opened yet — first non-blank, non-header row.
                if classification == RENTER_CLASS:
                    raise ValueError(
                        f"Renter row before any owner row at line {row_index}"
                    )
                # Other unclassified continuation rows are skipped silently
                # (e.g., a blank classification with a stray cell).
                continue

            # No name → skip (corner case: stray empty data).
            if not name:
                continue

            current.renters.append(
                ParsedTenant(
                    name=name,
                    phone_raw=phone,
                    classification=classification or RENTER_CLASS,
                )
            )

        return apartments
    finally:
        wb.close()


def parse_monthly_amounts(path: Path) -> dict[str, Decimal]:
    """Convenience: return ``{apartment_label: monthly_amount}``.

    Apartments whose row had no ``גובה תשלום`` value are skipped — only
    apartments with a non-None ``monthly_amount`` appear in the result.
    """
    return {
        apt.apartment_label: apt.monthly_amount
        for apt in parse_tenants_xlsx(path)
        if apt.monthly_amount is not None
    }
